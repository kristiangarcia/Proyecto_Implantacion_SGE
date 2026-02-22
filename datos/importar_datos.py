# -*- coding: utf-8 -*-
"""
Script de importación de datos a Odoo via XML-RPC.
Importa clientes, productos y empleados desde los archivos Excel/CSV proporcionados.
Proyecto de Implantación SGE - NexusTech Solutions S.L.
"""
import xmlrpc.client
import openpyxl
import csv
import os
import base64
import sys

# =============================================
# CONFIGURACIÓN DE CONEXIÓN
# =============================================
URL = 'http://localhost:8069'
BD = 'nexustech'
USUARIO = 'admin'
CLAVE = 'admin'

# Ruta a los datos de importación
RUTA_DATOS = r'C:\Users\Kristian\Downloads\Datos para importar-PROYECTO'

# =============================================
# CONEXIÓN A ODOO
# =============================================
def conectar():
    """Establece conexión con Odoo y devuelve uid y proxy de modelos."""
    comun = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/common')
    uid = comun.authenticate(BD, USUARIO, CLAVE, {})
    if not uid:
        print('ERROR: No se pudo autenticar con Odoo.')
        sys.exit(1)
    modelos = xmlrpc.client.ServerProxy(f'{URL}/xmlrpc/2/object')
    print(f'Conectado a Odoo (UID: {uid})')
    return uid, modelos


def ejecutar(modelos, uid, modelo, metodo, *args, **kwargs):
    """Ejecuta una operación en Odoo."""
    return modelos.execute_kw(BD, uid, CLAVE, modelo, metodo, *args, **kwargs)


# =============================================
# IMPORTAR CLIENTES
# =============================================
def importar_clientes(uid, modelos):
    """Importa clientes desde 1-1-b-ListadoClientes.xlsx"""
    archivo = os.path.join(RUTA_DATOS, '1-1-b-ListadoClientes.xlsx')
    if not os.path.exists(archivo):
        print(f'  Archivo no encontrado: {archivo}')
        return

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    # Leer cabeceras
    cabeceras = [cell.value for cell in ws[1]]
    print(f'  Cabeceras encontradas: {cabeceras}')

    contador = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos = dict(zip(cabeceras, fila))
        if not datos.get('Name'):
            continue

        valores = {
            'name': datos.get('Name', ''),
            'ref': datos.get('CLIENTE', ''),
            'vat': datos.get('Tax Id', '') if datos.get('Tax Id') else False,
            'email': datos.get('Main Email', '') if datos.get('Main Email') else False,
            'street': datos.get('Direccion', '') if datos.get('Direccion') else False,
            'city': datos.get('Ciudad', '') if datos.get('Ciudad') else False,
            'zip': str(datos.get('CP', '')) if datos.get('CP') else False,
            'customer_rank': 1,
            'is_company': True,
        }

        # Comprobar si ya existe
        existentes = ejecutar(modelos, uid, 'res.partner', 'search', [[['name', '=', valores['name']]]])
        if not existentes:
            ejecutar(modelos, uid, 'res.partner', 'create', [valores])
            contador += 1

    print(f'  Clientes importados: {contador}')


# =============================================
# IMPORTAR PRODUCTOS (SOFTWARE)
# =============================================
def importar_productos_software(uid, modelos):
    """Importa productos de software desde 4-3-d-ProductosSoftware.xlsx"""
    archivo = os.path.join(RUTA_DATOS, '4-3-d-ProductosSoftware.xlsx')
    if not os.path.exists(archivo):
        print(f'  Archivo no encontrado: {archivo}')
        return

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    cabeceras = [cell.value for cell in ws[1]]
    print(f'  Cabeceras: {cabeceras}')

    contador = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos = dict(zip(cabeceras, fila))
        nombre = datos.get('Nombre') or datos.get('nombre') or datos.get('Name')
        if not nombre:
            continue

        valores = {
            'name': str(nombre),
            'type': 'service',
            'sale_ok': True,
            'purchase_ok': True,
            'list_price': float(datos.get('Precio', 0) or datos.get('precio', 0) or 0),
        }

        # Añadir descripción si existe
        desc = datos.get('Descripción') or datos.get('descripcion') or datos.get('Description')
        if desc:
            valores['description'] = str(desc)

        existentes = ejecutar(modelos, uid, 'product.template', 'search', [[['name', '=', valores['name']]]])
        if not existentes:
            ejecutar(modelos, uid, 'product.template', 'create', [valores])
            contador += 1

    print(f'  Productos de software importados: {contador}')


# =============================================
# IMPORTAR PRODUCTOS (HARDWARE)
# =============================================
def importar_productos_hardware(uid, modelos):
    """Importa productos de hardware desde ProducHard.xlsx"""
    archivo = os.path.join(RUTA_DATOS, 'ProducHard.xlsx')
    if not os.path.exists(archivo):
        print(f'  Archivo no encontrado: {archivo}')
        return

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    cabeceras = [cell.value for cell in ws[1]]
    print(f'  Cabeceras: {cabeceras}')

    contador = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos = dict(zip(cabeceras, fila))
        nombre = datos.get('Nombre') or datos.get('nombre') or datos.get('Name') or datos.get('name')
        if not nombre:
            continue

        valores = {
            'name': str(nombre),
            'type': 'consu',
            'sale_ok': True,
            'purchase_ok': True,
            'list_price': float(datos.get('Precio', 0) or datos.get('precio', 0) or datos.get('Price', 0) or 0),
        }

        desc = datos.get('Descripción') or datos.get('descripcion') or datos.get('Description')
        if desc:
            valores['description'] = str(desc)

        existentes = ejecutar(modelos, uid, 'product.template', 'search', [[['name', '=', valores['name']]]])
        if not existentes:
            ejecutar(modelos, uid, 'product.template', 'create', [valores])
            contador += 1

    print(f'  Productos de hardware importados: {contador}')


# =============================================
# IMPORTAR SERVICIOS
# =============================================
def importar_servicios(uid, modelos):
    """Importa servicios desde 4-3-d-Servicios.xlsx"""
    archivo = os.path.join(RUTA_DATOS, '4-3-d-Servicios.xlsx')
    if not os.path.exists(archivo):
        print(f'  Archivo no encontrado: {archivo}')
        return

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    cabeceras = [cell.value for cell in ws[1]]
    print(f'  Cabeceras: {cabeceras}')

    contador = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos = dict(zip(cabeceras, fila))
        nombre = datos.get('Nombre') or datos.get('nombre') or datos.get('Name')
        if not nombre:
            continue

        valores = {
            'name': str(nombre),
            'type': 'service',
            'sale_ok': True,
            'purchase_ok': False,
            'list_price': float(datos.get('Precio', 0) or datos.get('precio', 0) or 0),
        }

        desc = datos.get('Descripción') or datos.get('descripcion') or datos.get('Description')
        if desc:
            valores['description'] = str(desc)

        existentes = ejecutar(modelos, uid, 'product.template', 'search', [[['name', '=', valores['name']]]])
        if not existentes:
            ejecutar(modelos, uid, 'product.template', 'create', [valores])
            contador += 1

    print(f'  Servicios importados: {contador}')


# =============================================
# IMPORTAR EMPLEADOS / USUARIOS
# =============================================
def importar_empleados(uid, modelos):
    """Importa empleados desde Lista_empleados_usuario.xlsx"""
    archivo = os.path.join(RUTA_DATOS, 'Lista_empleados_usuario.xlsx')
    if not os.path.exists(archivo):
        print(f'  Archivo no encontrado: {archivo}')
        return

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    cabeceras = [cell.value for cell in ws[1]]
    print(f'  Cabeceras: {cabeceras}')

    contador = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos = dict(zip(cabeceras, fila))
        # Intentar obtener nombre del empleado
        nombre = None
        for clave in ['Nombre', 'nombre', 'Name', 'name', 'Empleado', 'empleado']:
            if datos.get(clave):
                nombre = str(datos[clave])
                break

        if not nombre:
            # Usar primera columna con valor
            for val in fila:
                if val and str(val).strip():
                    nombre = str(val).strip()
                    break

        if not nombre:
            continue

        valores = {
            'name': nombre,
        }

        # Intentar añadir departamento
        depto = datos.get('Departamento') or datos.get('departamento') or datos.get('Department')
        if depto:
            # Buscar o crear departamento
            depto_ids = ejecutar(modelos, uid, 'hr.department', 'search', [[['name', '=', str(depto)]]])
            if not depto_ids:
                depto_id = ejecutar(modelos, uid, 'hr.department', 'create', [{'name': str(depto)}])
                depto_ids = [depto_id]
            valores['department_id'] = depto_ids[0]

        # Intentar añadir puesto
        puesto = datos.get('Puesto') or datos.get('puesto') or datos.get('Job') or datos.get('Cargo')
        if puesto:
            puesto_ids = ejecutar(modelos, uid, 'hr.job', 'search', [[['name', '=', str(puesto)]]])
            if not puesto_ids:
                puesto_id = ejecutar(modelos, uid, 'hr.job', 'create', [{'name': str(puesto)}])
                puesto_ids = [puesto_id]
            valores['job_id'] = puesto_ids[0]

        existentes = ejecutar(modelos, uid, 'hr.employee', 'search', [[['name', '=', valores['name']]]])
        if not existentes:
            ejecutar(modelos, uid, 'hr.employee', 'create', [valores])
            contador += 1

    print(f'  Empleados importados: {contador}')


# =============================================
# CONFIGURAR EMPRESA
# =============================================
def configurar_empresa(uid, modelos):
    """Configura los datos de la empresa NexusTech Solutions S.L."""
    # Obtener la empresa principal (id=1)
    empresa_ids = ejecutar(modelos, uid, 'res.company', 'search', [[]])
    if empresa_ids:
        ejecutar(modelos, uid, 'res.company', 'write', [empresa_ids[:1], {
            'name': 'NexusTech Solutions S.L.',
            'street': 'Calle Tecnología, 42',
            'city': 'Madrid',
            'zip': '28001',
            'country_id': 67,  # España
            'phone': '+34 912 345 678',
            'email': 'info@nexustech.es',
            'website': 'https://www.nexustech.es',
            'vat': 'ESB12345678',
            'currency_id': 1,  # EUR
        }])
        print('  Empresa configurada: NexusTech Solutions S.L.')


# =============================================
# PRINCIPAL
# =============================================
def main():
    """Función principal que ejecuta todas las importaciones."""
    print('=' * 60)
    print('IMPORTACIÓN DE DATOS - NexusTech Solutions S.L.')
    print('=' * 60)

    uid, modelos = conectar()

    print('\n1. Configurando empresa...')
    configurar_empresa(uid, modelos)

    print('\n2. Importando clientes...')
    importar_clientes(uid, modelos)

    print('\n3. Importando productos de software...')
    importar_productos_software(uid, modelos)

    print('\n4. Importando productos de hardware...')
    importar_productos_hardware(uid, modelos)

    print('\n5. Importando servicios...')
    importar_servicios(uid, modelos)

    print('\n6. Importando empleados...')
    importar_empleados(uid, modelos)

    print('\n' + '=' * 60)
    print('IMPORTACIÓN COMPLETADA')
    print('=' * 60)


if __name__ == '__main__':
    main()
